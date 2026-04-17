import streamlit as st
import cv2
import os
import gc
import numpy as np
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import torch
import torch.nn.functional as F
import torchvision.models as tv_models
import torchvision.transforms as T
from PIL import Image
from collections import defaultdict
from pathlib import Path
import tempfile
import datetime
import io
import time
import subprocess
import shutil

# ─── Page Config ───────────────────────────────────────────────────────────────
st.set_page_config(
    page_title="Person Re-ID · FastReID Dashboard",
    page_icon="🎯",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ─── Custom CSS ────────────────────────────────────────────────────────────────
st.markdown("""
<style>
  /* Overall background */
  .stApp { background: #0f1117; color: #e0e0e0; }

  /* Sidebar */
  section[data-testid="stSidebar"] {
    background: #161b27;
    border-right: 1px solid #2a3045;
  }

  /* Cards */
  .metric-card {
    background: linear-gradient(135deg, #1c2333, #1a2540);
    border: 1px solid #2e3f5c;
    border-radius: 12px;
    padding: 18px 22px;
    text-align: center;
    box-shadow: 0 4px 16px rgba(0,0,0,0.4);
  }
  .metric-card .label {
    font-size: 0.75rem;
    text-transform: uppercase;
    letter-spacing: 1px;
    color: #7a8fb5;
    margin-bottom: 6px;
  }
  .metric-card .value {
    font-size: 2rem;
    font-weight: 700;
    color: #4fc3f7;
  }
  .metric-card .value.green { color: #66bb6a; }
  .metric-card .value.amber { color: #ffa726; }
  .metric-card .value.purple { color: #ab47bc; }

  /* Section headers */
  .section-title {
    font-size: 1.1rem;
    font-weight: 600;
    color: #90caf9;
    border-left: 3px solid #4fc3f7;
    padding-left: 10px;
    margin: 18px 0 10px 0;
  }

  /* Status badges */
  .badge {
    display: inline-block;
    padding: 3px 10px;
    border-radius: 20px;
    font-size: 0.75rem;
    font-weight: 600;
    letter-spacing: 0.5px;
  }
  .badge-green  { background: #1b3a2a; color: #66bb6a; border: 1px solid #2e6b4a; }
  .badge-blue   { background: #1a2c45; color: #4fc3f7; border: 1px solid #2e5070; }
  .badge-amber  { background: #3a2810; color: #ffa726; border: 1px solid #6b4a20; }
  .badge-red    { background: #3a1a1a; color: #ef5350; border: 1px solid #6b2a2a; }
  .badge-purple { background: #2a1b3a; color: #ab47bc; border: 1px solid #4a2b6b; }

  /* Table styling */
  .stDataFrame { border-radius: 10px; overflow: hidden; }
  .stDataFrame thead th {
    background: #1c2333 !important;
    color: #90caf9 !important;
  }

  /* Button */
  .stButton > button {
    background: linear-gradient(135deg, #1565c0, #0d47a1);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 24px;
    font-weight: 600;
    letter-spacing: 0.5px;
    transition: all 0.2s;
    width: 100%;
  }
  .stButton > button:hover {
    background: linear-gradient(135deg, #1976d2, #1565c0);
    box-shadow: 0 4px 14px rgba(21,101,192,0.5);
  }

  /* Run button special */
  .run-btn > button {
    background: linear-gradient(135deg, #00796b, #004d40) !important;
    font-size: 1.05rem;
    padding: 14px 28px !important;
  }
  .run-btn > button:hover {
    background: linear-gradient(135deg, #00897b, #00695c) !important;
    box-shadow: 0 4px 18px rgba(0,121,107,0.55) !important;
  }

  /* Download button */
  .stDownloadButton > button {
    background: linear-gradient(135deg, #6a1b9a, #4a148c);
    color: white;
    border: none;
    border-radius: 8px;
    padding: 10px 22px;
    font-weight: 600;
    width: 100%;
  }

  /* Expander */
  .streamlit-expanderHeader {
    background: #1c2333;
    border-radius: 8px;
    color: #90caf9;
  }

  /* Slider */
  .stSlider > div > div > div { background: #4fc3f7; }

  /* Info / warning / success boxes */
  .stAlert { border-radius: 8px; }

  /* Hide default Streamlit branding */
  #MainMenu, footer { visibility: hidden; }

  /* Tab styling */
  .stTabs [data-baseweb="tab-list"] {
    background: #161b27;
    border-radius: 8px;
    padding: 4px;
    gap: 4px;
  }
  .stTabs [data-baseweb="tab"] {
    background: transparent;
    border-radius: 6px;
    color: #7a8fb5;
    font-weight: 500;
  }
  .stTabs [aria-selected="true"] {
    background: #1c2d4a !important;
    color: #4fc3f7 !important;
  }

  /* Log box */
  .log-box {
    background: #0d1117;
    border: 1px solid #2a3045;
    border-radius: 8px;
    padding: 14px 16px;
    font-family: monospace;
    font-size: 0.82rem;
    color: #c0c8d8;
    max-height: 300px;
    overflow-y: auto;
    line-height: 1.7;
  }

  /* Pipeline step indicator */
  .step-active   { color: #4fc3f7; font-weight: 700; }
  .step-done     { color: #66bb6a; }
  .step-pending  { color: #4a5568; }
</style>
""", unsafe_allow_html=True)


# ═══════════════════════════════════════════════════════════════════════════════
#  CONSTANTS (from notebook Cell 4)
# ═══════════════════════════════════════════════════════════════════════════════
EMBEDDING_DIM         = 2048
MAX_EMBS_PER_PERSON   = 100
EMA_ALPHA             = 0.90
DIVERSITY_MIN_DIST    = 0.08
RECENTLY_LOST_WINDOW  = 12000
RECENTLY_LOST_THR     = 0.60
RECENTLY_LOST_TOP_K   = 20
STABLE_ZONE_FRAMES    = 4
EVENT_COOLDOWN        = 60
MERGE_INTERVAL        = 300
MERGE_SIM_THRESHOLD   = 0.90
IDENTITY_UPGRADE_GAP  = 0.25
MIN_CROP_H            = 90
MIN_CROP_W            = 45
MIN_CROP_AREA         = 5000
CROP_ASPECT_MIN       = 1.5
CROP_ASPECT_MAX       = 5.0
MIN_DET_CONF_STORE    = 0.55
MIN_DET_CONF_EMA      = 0.40
GHOST_RATIO_TRIGGER   = 1.5
POSITION_PENALTY_DIST = 220
POSITION_PENALTY_W    = 0.82
DETECTION_NMS_IOU     = 0.50
CROWD_THRESHOLD       = 5
REEVAL_INTERVAL       = 25


# ═══════════════════════════════════════════════════════════════════════════════
#  CACHED MODEL LOADING
# ═══════════════════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner=False)
def load_reid_model():
    device   = 'cuda' if torch.cuda.is_available() else 'cpu'
    use_fp16 = torch.cuda.is_available()

    class FastReIDExtractor(torch.nn.Module):
        def __init__(self, pretrained=True):
            super().__init__()
            backbone = tv_models.resnet50(
                weights=tv_models.ResNet50_Weights.IMAGENET1K_V2 if pretrained else None)
            self.features = torch.nn.Sequential(*list(backbone.children())[:-1])
            self.eval()

        def forward(self, x):
            with torch.no_grad():
                feat = self.features(x)
                feat = feat.flatten(1)
                feat = F.normalize(feat, p=2, dim=1)
            return feat

    model = FastReIDExtractor(pretrained=True).to(device)
    model.eval()
    if use_fp16:
        model = model.half()
    return model, device, use_fp16


@st.cache_resource(show_spinner=False)
def load_yolo_models():
    try:
        from ultralytics import YOLO
        fast  = YOLO('yolov8n.pt')
        final = YOLO('yolov8m.pt')
        return fast, final, True
    except Exception as e:
        return None, None, False


# ═══════════════════════════════════════════════════════════════════════════════
#  UTILITY FUNCTIONS
# ═══════════════════════════════════════════════════════════════════════════════
def get_transforms():
    return T.Compose([
        T.Resize((256, 128)),
        T.ToTensor(),
        T.Normalize(mean=[0.485, 0.456, 0.406], std=[0.229, 0.224, 0.225]),
    ])


def nms_detections(dets, iou_thr=DETECTION_NMS_IOU):
    if not dets:
        return []
    dets = sorted(dets, key=lambda d: d[4], reverse=True)
    keep, suppressed = [], set()
    for i, d in enumerate(dets):
        if i in suppressed:
            continue
        keep.append(d)
        x1a, y1a, x2a, y2a = d[0], d[1], d[2], d[3]
        aA = max(0, x2a - x1a) * max(0, y2a - y1a)
        for j in range(i + 1, len(dets)):
            if j in suppressed:
                continue
            x1b, y1b, x2b, y2b = dets[j][0], dets[j][1], dets[j][2], dets[j][3]
            ix1 = max(x1a, x1b); iy1 = max(y1a, y1b)
            ix2 = min(x2a, x2b); iy2 = min(y2a, y2b)
            inter = max(0, ix2 - ix1) * max(0, iy2 - iy1)
            if inter == 0:
                continue
            aB = max(0, x2b - x1b) * max(0, y2b - y1b)
            if inter / (aA + aB - inter + 1e-6) >= iou_thr:
                suppressed.add(j)
    return keep


def crop_quality_ok(crop):
    h, w = crop.shape[:2]
    if h < MIN_CROP_H or w < MIN_CROP_W or h * w < MIN_CROP_AREA:
        return False
    aspect = h / (w + 1e-6)
    return CROP_ASPECT_MIN <= aspect <= CROP_ASPECT_MAX


def run_fastreid(crops_bgr, reid_model, device, use_fp16, transform):
    tensors = []
    for crop in crops_bgr:
        pil = Image.fromarray(cv2.cvtColor(crop, cv2.COLOR_BGR2RGB))
        tensors.append(transform(pil))
    batch = torch.stack(tensors, dim=0).to(device)
    if use_fp16:
        batch = batch.half()
    with torch.no_grad():
        feats = reid_model(batch).float()
    return feats.cpu().numpy().astype(np.float32)


def extract_embedding_fast(crop, reid_model, device, use_fp16, transform):
    if not crop_quality_ok(crop):
        return None
    crops = [crop, cv2.flip(crop, 1)]
    embs  = run_fastreid(crops, reid_model, device, use_fp16, transform)
    fused = embs.mean(axis=0)
    norm  = np.linalg.norm(fused) + 1e-8
    return (fused / norm).astype(np.float32)


def extract_embedding_full(crop, reid_model, device, use_fp16, transform):
    if not crop_quality_ok(crop):
        return None
    h = crop.shape[0]
    upper = crop[:h//2, :]
    lower = crop[h//2:, :]
    batch_crops   = [crop, cv2.flip(crop, 1)]
    batch_weights = [1.0, 1.0]
    if crop_quality_ok(upper):
        batch_crops  += [upper, cv2.flip(upper, 1)]
        batch_weights += [0.65, 0.65]
    if crop_quality_ok(lower):
        batch_crops  += [lower, cv2.flip(lower, 1)]
        batch_weights += [0.65, 0.65]
    if crop.shape[0] > 100 and crop.shape[1] > 40:
        cx, cy = crop.shape[1]//2, crop.shape[0]//2
        for angle in [5, -5]:
            M   = cv2.getRotationMatrix2D((cx, cy), angle, 1.0)
            rot = cv2.warpAffine(crop, M, (crop.shape[1], crop.shape[0]))
            batch_crops.append(rot)
            batch_weights.append(0.5)
    embs  = run_fastreid(batch_crops, reid_model, device, use_fp16, transform)
    w     = np.array(batch_weights, dtype=np.float32)
    fused = (embs * w[:, None]).sum(axis=0) / w.sum()
    norm  = np.linalg.norm(fused) + 1e-8
    return (fused / norm).astype(np.float32)


def iou_fn(boxA, boxB):
    xA = max(boxA[0], boxB[0]); yA = max(boxA[1], boxB[1])
    xB = min(boxA[2], boxB[2]); yB = min(boxA[3], boxB[3])
    inter = max(0, xB - xA) * max(0, yB - yA)
    if inter == 0:
        return 0.0
    aA = (boxA[2]-boxA[0]) * (boxA[3]-boxA[1])
    aB = (boxB[2]-boxB[0]) * (boxB[3]-boxB[1])
    return inter / (aA + aB - inter + 1e-6)


class IoUTracker:
    def __init__(self, iou_threshold=0.35, max_lost=30):
        self.iou_threshold = iou_threshold
        self.max_lost = max_lost
        self._tracks = {}
        self._next_tid = 1

    def update(self, detections):
        if not detections:
            for tid in list(self._tracks):
                self._tracks[tid]['lost'] += 1
                if self._tracks[tid]['lost'] > self.max_lost:
                    del self._tracks[tid]
            return []
        track_ids   = list(self._tracks.keys())
        track_boxes = [self._tracks[t]['box'] for t in track_ids]
        matched_tracks, matched_dets = set(), set()
        results = []
        pairs = []
        for di, det in enumerate(detections):
            for ti, tbox in enumerate(track_boxes):
                sc = iou_fn(det[:4], tbox)
                if sc >= self.iou_threshold:
                    pairs.append((sc, di, ti))
        pairs.sort(reverse=True)
        for sc, di, ti in pairs:
            if di in matched_dets or ti in matched_tracks:
                continue
            tid = track_ids[ti]
            self._tracks[tid].update(box=detections[di][:4], conf=detections[di][4], lost=0)
            matched_dets.add(di); matched_tracks.add(ti)
            results.append((*detections[di][:4], detections[di][4], tid))
        for di, det in enumerate(detections):
            if di not in matched_dets:
                tid = self._next_tid; self._next_tid += 1
                self._tracks[tid] = {'box': det[:4], 'conf': det[4], 'lost': 0}
                results.append((*det[:4], det[4], tid))
        for ti, tid in enumerate(track_ids):
            if ti not in matched_tracks:
                self._tracks[tid]['lost'] += 1
                if self._tracks[tid]['lost'] > self.max_lost:
                    del self._tracks[tid]
        return results


try:
    import faiss
    FAISS_AVAILABLE = True
except ImportError:
    FAISS_AVAILABLE = False


class FAISSGallery:
    def __init__(self, dim):
        self.dim = dim
        if FAISS_AVAILABLE:
            cpu_idx = faiss.IndexFlatIP(dim)
            self._on_gpu = False
            if torch.cuda.is_available():
                try:
                    res = faiss.StandardGpuResources()
                    self.index = faiss.index_cpu_to_gpu(res, 0, cpu_idx)
                    self._on_gpu = True
                except Exception:
                    self.index = cpu_idx
            else:
                self.index = cpu_idx
        else:
            # Pure-numpy fallback
            self._vectors = []
            self._on_gpu = False
        self.id_map = []

    def add(self, emb, pid):
        if FAISS_AVAILABLE:
            self.index.add(emb.reshape(1, -1).astype(np.float32))
        else:
            self._vectors.append(emb.copy())
        self.id_map.append(pid)

    def search(self, query, k=20):
        if FAISS_AVAILABLE:
            if self.index.ntotal == 0:
                return np.array([]), np.array([])
            k = min(k, self.index.ntotal)
            scores, idxs = self.index.search(query.reshape(1,-1).astype(np.float32), k)
            return scores[0], np.array([self.id_map[i] for i in idxs[0]])
        else:
            if not self._vectors:
                return np.array([]), np.array([])
            mat = np.stack(self._vectors)
            sims = (mat @ query.reshape(-1)).astype(np.float32)
            k = min(k, len(sims))
            top_idx = np.argsort(sims)[::-1][:k]
            return sims[top_idx], np.array([self.id_map[i] for i in top_idx])

    def rebuild(self, person_database):
        if FAISS_AVAILABLE:
            cpu_idx = faiss.IndexFlatIP(self.dim)
            new_ids = []
            for pid, rec in person_database.items():
                emb = rec.get('ema_embedding')
                if emb is not None:
                    cpu_idx.add(emb.reshape(1,-1).astype(np.float32))
                    new_ids.append(pid)
            if torch.cuda.is_available() and self._on_gpu:
                try:
                    res = faiss.StandardGpuResources()
                    self.index = faiss.index_cpu_to_gpu(res, 0, cpu_idx)
                except Exception:
                    self.index = cpu_idx
            else:
                self.index = cpu_idx
            self.id_map = new_ids
        else:
            self._vectors = []
            self.id_map   = []
            for pid, rec in person_database.items():
                emb = rec.get('ema_embedding')
                if emb is not None:
                    self._vectors.append(emb.copy())
                    self.id_map.append(pid)


def position_multiplier(pid, query_centroid, person_database):
    if query_centroid is None: return 1.0
    rec = person_database.get(pid)
    if rec is None: return 1.0
    last_xy = rec.get('last_centroid')
    if last_xy is None: return 1.0
    dist = ((query_centroid[0]-last_xy[0])**2 + (query_centroid[1]-last_xy[1])**2)**0.5
    return 1.0 if dist < POSITION_PENALTY_DIST else POSITION_PENALTY_W


def match_identity(query_emb, gallery, person_database, recently_lost,
                   sim_thr, occupied_pids, n_active, centroid=None):
    if (FAISS_AVAILABLE and gallery.index.ntotal == 0) or \
       (not FAISS_AVAILABLE and not gallery._vectors):
        return None, 0.0
    scores, pids = gallery.search(query_emb, k=RECENTLY_LOST_TOP_K)
    if len(scores) == 0:
        return None, 0.0
    ghost_ratio   = len(occupied_pids) / max(n_active, 1)
    effective_thr = sim_thr * (POSITION_PENALTY_W if ghost_ratio > GHOST_RATIO_TRIGGER else 1.0)
    pid_scores = defaultdict(list)
    for sc, pid in zip(scores, pids):
        pid_scores[pid].append(sc)
    best_pid, best_score = None, 0.0
    for pid, sc_list in pid_scores.items():
        agg = float(np.mean(sc_list)) * position_multiplier(pid, centroid, person_database)
        if pid in recently_lost:
            agg = max(agg, recently_lost[pid] * RECENTLY_LOST_THR)
        if agg > best_score:
            best_score = agg; best_pid = pid
    if best_score < effective_thr:
        return None, 0.0
    return best_pid, best_score


def register_new_person(emb, frame_num, person_database, gallery, next_pid,
                        max_persons_expected, centroid=None):
    if len(person_database) >= max_persons_expected:
        scores, pids = gallery.search(emb, k=1)
        if len(scores) > 0:
            return int(pids[0])
    pid = next_pid[0]; next_pid[0] += 1
    person_database[pid] = {
        'embeddings':    [emb.copy()],
        'ema_embedding': emb.copy(),
        'first_seen':    frame_num,
        'last_seen':     frame_num,
        'visit_count':   0,
        'last_centroid': centroid,
    }
    gallery.add(emb, pid)
    return pid


def update_person(pid, emb, frame_num, person_database, gallery,
                  det_conf=1.0, match_score=1.0, centroid=None):
    rec = person_database.get(pid)
    if rec is None: return
    rec['last_seen'] = frame_num
    if centroid is not None:
        rec['last_centroid'] = centroid
    if det_conf < MIN_DET_CONF_STORE:
        return
    if rec['embeddings']:
        existing = np.stack(rec['embeddings'])
        if (existing @ emb).max() > 1.0 - DIVERSITY_MIN_DIST:
            if det_conf >= MIN_DET_CONF_EMA:
                alpha = EMA_ALPHA * match_score
                ema   = alpha * rec['ema_embedding'] + (1 - alpha) * emb
                ema  /= (np.linalg.norm(ema) + 1e-8)
                rec['ema_embedding'] = ema
            return
    rec['embeddings'].append(emb.copy())
    if len(rec['embeddings']) > MAX_EMBS_PER_PERSON:
        rec['embeddings'].pop(0)
    if det_conf >= MIN_DET_CONF_EMA:
        alpha = EMA_ALPHA * match_score
        ema   = alpha * rec['ema_embedding'] + (1 - alpha) * emb
        ema  /= (np.linalg.norm(ema) + 1e-8)
        rec['ema_embedding'] = ema
        gallery.add(ema, pid)


def maybe_merge_identities(person_database, gallery, frame_num):
    if frame_num % MERGE_INTERVAL != 0: return
    pids = list(person_database.keys())
    if len(pids) < 2: return
    embs = np.stack([person_database[p]['ema_embedding'] for p in pids])
    sim  = embs @ embs.T
    to_merge = [(pids[i], pids[j])
                for i in range(len(pids))
                for j in range(i+1, len(pids))
                if sim[i,j] >= MERGE_SIM_THRESHOLD]
    for pa, pb in to_merge:
        if pa not in person_database or pb not in person_database: continue
        ra, rb = person_database[pa], person_database[pb]
        ema  = 0.5 * ra['ema_embedding'] + 0.5 * rb['ema_embedding']
        ema /= (np.linalg.norm(ema) + 1e-8)
        ra['ema_embedding'] = ema
        ra['embeddings']    = (ra['embeddings'] + rb['embeddings'])[-MAX_EMBS_PER_PERSON:]
        ra['visit_count']  += rb.get('visit_count', 0)
        del person_database[pb]
    if to_merge:
        gallery.rebuild(person_database)


def pid_color(pid):
    rng = np.random.RandomState(pid * 137 + 7)
    return tuple(int(c) for c in rng.randint(60, 220, 3))


def run_pipeline(video_path, cfg, yolo_model, reid_model, device, use_fp16,
                 progress_bar=None, status_text=None, log_container=None):
    CONF_THR  = cfg.get('conf_thr', 0.45)
    SIM_THR   = cfg.get('sim_thr',  0.82)
    MIN_TF    = cfg.get('min_track_frames', 5)
    ENTRY_R   = cfg.get('entry_y',  0.35)
    EXIT_R    = cfg.get('exit_y',   0.65)
    SKIP      = cfg.get('frame_skip', 3)
    IOU_THR   = cfg.get('iou_thr',  0.35)
    MAX_LOST  = cfg.get('max_lost', 30)
    MAX_PERS  = cfg.get('max_persons_expected', 8)
    fast_tta  = cfg.get('fast_tta', False)

    transform = get_transforms()
    tracker   = IoUTracker(iou_threshold=IOU_THR, max_lost=MAX_LOST)

    person_database   = {}
    next_pid          = [1]
    event_log         = []
    gallery           = FAISSGallery(EMBEDDING_DIM)
    recently_lost     = {}

    track_emb_buffer  = defaultdict(list)
    track_identity    = {}
    track_similarity  = {}
    track_confidence  = {}
    track_last_reeval = {}
    track_zone        = {}
    track_zone_stable = {}
    track_zone_cnt    = {}

    pid_last_event_frame     = defaultdict(lambda: defaultdict(lambda: -9999))
    pid_last_confirmed_event = {}
    track_counted_entry      = set()
    log_lines                = []

    embed_fn = (lambda c: extract_embedding_fast(c, reid_model, device, use_fp16, transform)) \
               if fast_tta else \
               (lambda c: extract_embedding_full(c, reid_model, device, use_fp16, transform))

    cap   = cv2.VideoCapture(video_path)
    W     = int(cap.get(cv2.CAP_PROP_FRAME_WIDTH))
    H     = int(cap.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps   = cap.get(cv2.CAP_PROP_FPS) or 30.0
    total = int(cap.get(cv2.CAP_PROP_FRAME_COUNT))
    cap.release()

    entry_y  = int(H * ENTRY_R)
    exit_y   = int(H * EXIT_R)
    out_path = str(Path(video_path).with_suffix('')) + '_reid_out.mp4'
    writer   = cv2.VideoWriter(out_path, cv2.VideoWriter_fourcc(*'mp4v'), fps, (W, H))

    def zone_of(cy):
        if cy < entry_y: return 'above'
        if cy > exit_y:  return 'below'
        return 'between'

    def fire_event(pid, event, fn):
        last = pid_last_confirmed_event.get(pid)
        if event == 'EXITED' and last != 'ENTERED': return False
        if fn - pid_last_event_frame[pid][event] < EVENT_COOLDOWN: return False
        pid_last_event_frame[pid][event] = fn
        pid_last_confirmed_event[pid]    = event
        if event == 'ENTERED':
            track_counted_entry.add(pid)
            person_database[pid]['visit_count'] = person_database[pid].get('visit_count', 0) + 1
        event_log.append({'frame': fn, 'person_id': pid, 'event': event})
        log_lines.append(('🟢' if event == 'ENTERED' else '🔴') +
                         f' {event} — Person #{pid} | Frame {fn}')
        return True

    def check_zone(tid, pid, cy, fn):
        if pid is None: return
        nz = zone_of(cy)
        if track_zone.get(tid) != nz:
            track_zone[tid] = nz; track_zone_cnt[tid] = 1
        else:
            track_zone_cnt[tid] = track_zone_cnt.get(tid, 0) + 1
        if track_zone_cnt.get(tid, 0) < STABLE_ZONE_FRAMES: return
        stable = track_zone_stable.get(tid)
        event  = None
        if stable is None:
            if pid not in track_counted_entry: event = 'ENTERED'
        else:
            if stable == 'above' and nz in ('between','below'):    event = 'ENTERED'
            elif stable in ('between','below') and nz == 'above':  event = 'EXITED'
            elif stable == 'between' and nz == 'below':            event = 'EXITED'
        track_zone_stable[tid] = nz
        if event: fire_event(pid, event, fn)

    def get_identity(tid, emb, fn, occupied_pids, det_conf, centroid, n_active):
        buf = track_emb_buffer[tid]
        buf.append((emb.copy(), det_conf))
        if len(buf) < MIN_TF:
            return None
        current_pid = track_identity.get(tid)
        last_reeval = track_last_reeval.get(tid, -9999)
        if current_pid is not None and (fn - last_reeval) < REEVAL_INTERVAL:
            update_person(current_pid, emb, fn, person_database, gallery,
                          det_conf=det_conf, match_score=track_confidence.get(tid, 1.0),
                          centroid=centroid)
            return current_pid
        recent    = buf[-MIN_TF:]
        best_emb  = max(recent, key=lambda x: x[1])[0]
        mean_emb  = np.stack([b[0] for b in recent]).mean(axis=0)
        mean_emb /= (np.linalg.norm(mean_emb) + 1e-8)
        query_emb = 0.35 * best_emb + 0.25 * mean_emb
        if current_pid is not None and current_pid in person_database:
            ema       = person_database[current_pid]['ema_embedding']
            query_emb = 0.40 * query_emb + 0.40 * ema
        query_emb /= (np.linalg.norm(query_emb) + 1e-8)
        track_last_reeval[tid] = fn
        pid, score = match_identity(
            query_emb, gallery, person_database, recently_lost,
            SIM_THR, occupied_pids, n_active, centroid)
        if current_pid is None:
            if pid is None:
                pid   = register_new_person(query_emb, fn, person_database,
                                            gallery, next_pid, MAX_PERS, centroid)
                score = 1.0
            recently_lost.pop(pid, None)
            track_identity[tid]   = pid
            track_similarity[tid] = score
            track_confidence[tid] = score
            update_person(pid, emb, fn, person_database, gallery,
                          det_conf=det_conf, match_score=score, centroid=centroid)
            return pid
        current_conf = track_confidence.get(tid, 0.0)
        if pid is not None and pid != current_pid and score > current_conf + IDENTITY_UPGRADE_GAP:
            track_identity[tid]   = pid
            track_similarity[tid] = score
            track_confidence[tid] = score
            recently_lost.pop(pid, None)
            update_person(pid, emb, fn, person_database, gallery,
                          det_conf=det_conf, match_score=score, centroid=centroid)
            return pid
        track_similarity[tid] = current_conf
        update_person(current_pid, emb, fn, person_database, gallery,
                      det_conf=det_conf, match_score=current_conf, centroid=centroid)
        return current_pid

    def handle_lost_tracks(active_tids, fn):
        for tid, pid in list(track_identity.items()):
            if tid not in active_tids and pid in person_database:
                age = fn - person_database[pid].get('last_seen', fn)
                if 0 < age <= RECENTLY_LOST_WINDOW:
                    recently_lost[pid] = track_confidence.get(tid, 0.5)

    cap  = cv2.VideoCapture(video_path)
    fn   = 0
    t0   = time.time()

    while True:
        ret, frame = cap.read()
        if not ret: break
        fn += 1

        if fn % SKIP != 0:
            writer.write(frame)
            continue

        # Progress update every 30 frames
        if progress_bar is not None and fn % 30 == 0:
            pct = fn / max(total, 1)
            progress_bar.progress(min(pct, 1.0))
            elapsed = time.time() - t0
            fps_proc = fn / max(elapsed, 0.1)
            eta = (total - fn) / max(fps_proc, 0.1)
            if status_text:
                status_text.markdown(
                    f"⚙️ Processing frame **{fn}/{total}** &nbsp;|&nbsp; "
                    f"Persons found: **{len(person_database)}** &nbsp;|&nbsp; "
                    f"ETA: **{eta:.0f}s**")

        results  = yolo_model(frame, classes=[0], conf=CONF_THR, verbose=False)
        raw_dets = []
        for r in results:
            for box in r.boxes:
                x1, y1, x2, y2 = map(int, box.xyxy[0].tolist())
                cf = float(box.conf[0])
                raw_dets.append((x1, y1, x2, y2, cf))

        det_list    = nms_detections(raw_dets, iou_thr=DETECTION_NMS_IOU)
        tracks      = tracker.update(det_list)
        active_tids = {t[5] for t in tracks}

        handle_lost_tracks(active_tids, fn)

        for pid in list(recently_lost):
            if pid in person_database:
                if fn - person_database[pid].get('last_seen', 0) > RECENTLY_LOST_WINDOW:
                    del recently_lost[pid]

        occupied_pids = {track_identity.get(t[5]) for t in tracks if track_identity.get(t[5]) is not None}
        n_active = len(tracks)

        maybe_merge_identities(person_database, gallery, fn)

        out_frame = frame.copy()
        cv2.line(out_frame, (0, entry_y), (W, entry_y), (0, 255, 0), 2)
        cv2.line(out_frame, (0, exit_y),  (W, exit_y),  (0, 0, 255), 2)

        for x1, y1, x2, y2, conf, tid in tracks:
            crop = frame[max(0,y1):y2, max(0,x1):x2]
            if crop.size == 0: continue
            emb = embed_fn(crop)
            cx  = (x1 + x2) // 2
            cy  = (y1 + y2) // 2
            pid = None
            if emb is not None:
                pid = get_identity(tid, emb, fn, occupied_pids, conf, (cx, cy), n_active)
            check_zone(tid, pid, cy, fn)
            color = pid_color(pid) if pid else (180, 180, 180)
            cv2.rectangle(out_frame, (x1, y1), (x2, y2), color, 2)
            label = f'P#{pid}' if pid else f'T#{tid}'
            if pid:
                label += f' {track_similarity.get(tid, 0.0):.2f}'
            cv2.putText(out_frame, label, (x1, y1-8),
                        cv2.FONT_HERSHEY_SIMPLEX, 0.65, color, 2)

        cv2.putText(out_frame,
                    f'Frame:{fn}  IDs:{len(person_database)}  TTA:{"fast" if fast_tta else "full"}',
                    (10, 28), cv2.FONT_HERSHEY_SIMPLEX, 0.75, (255,255,255), 2)
        writer.write(out_frame)

    cap.release()
    writer.release()

    entries  = sum(1 for e in event_log if e['event'] == 'ENTERED')
    exits    = sum(1 for e in event_log if e['event'] == 'EXITED')
    revisits = sum(1 for p in person_database.values() if p.get('visit_count', 0) > 1)
    n_ids    = len(person_database)

    metrics = {
        'total_persons':    n_ids,
        'total_entries':    entries,
        'total_exits':      exits,
        'revisit_persons':  revisits,
        'frames_processed': fn,
    }
    count_penalty = max(0, n_ids - MAX_PERS) * 1.5
    score = (revisits * 2.5 + min(entries, exits) * 0.4 +
             min(n_ids, MAX_PERS) * 0.3 - count_penalty)

    return {
        'metrics':     metrics,
        'score':       score,
        'event_log':   event_log,
        'database':    person_database,
        'log_lines':   log_lines,
        'output_path': out_path,
        'cfg':         cfg,
    }


def convert_to_h264(raw_path):
    out_path = raw_path.replace('.mp4', '_h264.mp4')
    if shutil.which('ffmpeg'):
        subprocess.run(
            ['ffmpeg', '-y', '-i', raw_path,
             '-vcodec', 'libx264', '-crf', '23', '-preset', 'fast', out_path],
            check=False, capture_output=True)
        if os.path.exists(out_path):
            return out_path
    return raw_path


def results_to_excel(result):
    db  = result['database']
    evs = result['event_log']

    rows = []
    for pid, rec in sorted(db.items()):
        rows.append({
            'person_id':   pid,
            'first_seen':  rec.get('first_seen', '?'),
            'last_seen':   rec.get('last_seen',  '?'),
            'visit_count': rec.get('visit_count', 0),
            'stored_embs': len(rec.get('embeddings', [])),
            're_identified': rec.get('visit_count', 0) > 1,
        })

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine='openpyxl') as writer_:
        pd.DataFrame(rows).to_excel(writer_, sheet_name='Persons', index=False)
        if evs:
            pd.DataFrame(evs).to_excel(writer_, sheet_name='Event Log', index=False)
        m = result['metrics']
        pd.DataFrame([
            {'metric': k, 'value': v} for k, v in m.items()
        ] + [{'metric': 'score', 'value': round(result['score'], 4)}]
        ).to_excel(writer_, sheet_name='Summary', index=False)
        pd.DataFrame([
            {'parameter': k, 'value': v} for k, v in result['cfg'].items()
        ]).to_excel(writer_, sheet_name='Config', index=False)

        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
        bold_font = Font(bold=True)
        for sn in writer_.sheets:
            ws = writer_.sheets[sn]
            for col in ws.columns:
                ml = max((len(str(c.value)) for c in col if c.value), default=8)
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(ml+4, 35)
            for cell in ws[1]:
                cell.font = bold_font

    buf.seek(0)
    return buf


# ═══════════════════════════════════════════════════════════════════════════════
#  UI: SIDEBAR
# ═══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("""
    <div style="text-align:center; padding: 16px 0 10px 0;">
      <div style="font-size:2rem;">🎯</div>
      <div style="font-size:1.15rem; font-weight:700; color:#4fc3f7; letter-spacing:0.5px;">
        Person Re-ID
      </div>
      <div style="font-size:0.72rem; color:#7a8fb5; margin-top:2px;">
        FastReID · FAISS · YOLOv8 · IoU Tracker
      </div>
    </div>
    <hr style="border-color:#2a3045; margin: 8px 0 18px 0;">
    """, unsafe_allow_html=True)

    # ── System Status ──────────────────────────────────────────────────────────
    device   = 'cuda' if torch.cuda.is_available() else 'cpu'
    gpu_name = torch.cuda.get_device_name(0) if torch.cuda.is_available() else 'N/A'
    st.markdown('<div class="section-title">⚙️ System</div>', unsafe_allow_html=True)

    col_a, col_b = st.columns(2)
    with col_a:
        badge = 'badge-green' if device == 'cuda' else 'badge-amber'
        label = 'GPU' if device == 'cuda' else 'CPU'
        st.markdown(f'<span class="badge {badge}">{label}</span>', unsafe_allow_html=True)
    with col_b:
        faiss_badge = 'badge-green' if FAISS_AVAILABLE else 'badge-amber'
        faiss_label = 'FAISS ✓' if FAISS_AVAILABLE else 'FAISS (numpy)'
        st.markdown(f'<span class="badge {faiss_badge}" style="font-size:0.65rem">{faiss_label}</span>',
                    unsafe_allow_html=True)

    if device == 'cuda':
        st.caption(f"🖥 {gpu_name}")

    st.markdown('<div class="section-title">📹 Detection</div>', unsafe_allow_html=True)
    yolo_choice = st.selectbox("YOLO Model", ['yolov8n.pt (fast)', 'yolov8m.pt (accurate)'],
                               index=1, help="n=fast, m=accurate")
    conf_thr  = st.slider("Confidence Threshold", 0.20, 0.80, 0.45, 0.05,
                           help="Higher → fewer false positives / ghost detections")
    iou_thr   = st.slider("Tracker IoU Threshold", 0.20, 0.65, 0.35, 0.05,
                           help="Min IoU to associate detection with an existing track")
    max_lost  = st.slider("Max Lost Frames", 5, 80, 30, 5,
                           help="Frames before a track is dropped")

    st.markdown('<div class="section-title">🧠 Re-ID</div>', unsafe_allow_html=True)
    sim_thr         = st.slider("Similarity Threshold", 0.60, 0.98, 0.82, 0.02,
                                 help="Cosine similarity needed to match an existing identity")
    min_track_frames = st.slider("Min Track Frames", 2, 15, 5, 1,
                                  help="Buffer length before assigning an ID")

    st.markdown('<div class="section-title">🎬 Pipeline</div>', unsafe_allow_html=True)
    frame_skip = st.slider("Frame Skip", 1, 8, 3, 1,
                            help="Process every Nth frame (higher = faster)")
    tta_mode   = st.selectbox("TTA Mode", ['Full (10 augments)', 'Fast (2 augments — flip only)'],
                               index=0, help="Test-time augmentation")
    entry_y    = st.slider("Entry Line (%)", 0.10, 0.50, 0.35, 0.05,
                            help="Entry zone upper boundary (fraction of frame height)")
    exit_y     = st.slider("Exit Line (%)", 0.50, 0.90, 0.65, 0.05,
                            help="Exit zone lower boundary (fraction of frame height)")
    max_persons = st.slider("Max Persons Expected", 2, 30, 8, 1,
                             help="~2× real number of people in scene; prevents phantom IDs")

    st.markdown("---")
    st.caption("**Tip:** Set Max Persons Expected to ~2× the real count in your video.")


# ═══════════════════════════════════════════════════════════════════════════════
#  UI: MAIN CONTENT
# ═══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="padding: 10px 0 20px 0;">
  <h1 style="font-size:1.8rem; font-weight:800; color:#e0e0e0; margin:0;">
    Person Re-Identification Dashboard
  </h1>
  <p style="color:#7a8fb5; margin:4px 0 0 0; font-size:0.9rem;">
    YOLOv8 → IoU Tracker → FastReID (ResNet-50) → FAISS Re-ID
  </p>
</div>
""", unsafe_allow_html=True)

tab_run, tab_results, tab_analytics, tab_download = st.tabs([
    "▶  Run Pipeline", "📊  Results & Events", "📈  Analytics", "💾  Download"])


# ─────────────────────────────────────────────────────────────────────────────
# TAB 1 — RUN PIPELINE
# ─────────────────────────────────────────────────────────────────────────────
with tab_run:
    col_upload, col_cfg = st.columns([1.5, 1], gap="large")

    with col_upload:
        st.markdown('<div class="section-title">📁 Upload Video</div>', unsafe_allow_html=True)
        video_file = st.file_uploader(
            "Drop a video here (.mp4, .avi, .mov, .mkv)",
            type=['mp4', 'avi', 'mov', 'mkv'],
            label_visibility='collapsed')

        if video_file:
            st.video(video_file)
            st.markdown(f"""
            <div style="background:#1c2333; border:1px solid #2e3f5c; border-radius:8px;
                        padding:12px 16px; margin-top:10px; font-size:0.83rem; color:#a0b4cc;">
              📄 <b>{video_file.name}</b> &nbsp;·&nbsp;
              {video_file.size / 1024 / 1024:.1f} MB
            </div>""", unsafe_allow_html=True)

    with col_cfg:
        st.markdown('<div class="section-title">🔧 Active Config</div>', unsafe_allow_html=True)
        cfg_preview = {
            'conf_thr':           conf_thr,
            'sim_thr':            sim_thr,
            'iou_thr':            iou_thr,
            'max_lost':           max_lost,
            'min_track_frames':   min_track_frames,
            'frame_skip':         frame_skip,
            'entry_y':            entry_y,
            'exit_y':             exit_y,
            'max_persons_expected': max_persons,
            'tta':                tta_mode.split('(')[0].strip(),
            'yolo_model':         yolo_choice.split('(')[0].strip(),
        }
        for k, v in cfg_preview.items():
            st.markdown(
                f'<div style="display:flex; justify-content:space-between; '
                f'padding:5px 0; border-bottom:1px solid #1e2a3a; font-size:0.83rem;">'
                f'<span style="color:#7a8fb5;">{k}</span>'
                f'<span style="color:#e0e0e0; font-weight:600;">{v}</span></div>',
                unsafe_allow_html=True)

    st.markdown("---")

    # ── Run button ─────────────────────────────────────────────────────────────
    run_col, _ = st.columns([1, 2])
    with run_col:
        st.markdown('<div class="run-btn">', unsafe_allow_html=True)
        run_btn = st.button("🚀  Run Re-ID Pipeline", use_container_width=True)
        st.markdown('</div>', unsafe_allow_html=True)

    if run_btn:
        if video_file is None:
            st.error("⚠️ Please upload a video first.")
        else:
            # Save upload to temp file
            suffix = Path(video_file.name).suffix
            with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
                tmp.write(video_file.read())
                tmp_path = tmp.name

            # Build config
            cfg = {
                'conf_thr':             conf_thr,
                'sim_thr':              sim_thr,
                'iou_thr':              iou_thr,
                'max_lost':             max_lost,
                'min_track_frames':     min_track_frames,
                'frame_skip':           frame_skip,
                'entry_y':              entry_y,
                'exit_y':               exit_y,
                'max_persons_expected': max_persons,
                'fast_tta':             'Fast' in tta_mode,
                'yolo_model':           yolo_choice.split('(')[0].strip(),
            }

            # Load models
            with st.spinner("Loading models…"):
                reid_model, dev, fp16 = load_reid_model()
                yolo_fast, yolo_final, yolo_ok = load_yolo_models()

            if not yolo_ok:
                st.error("❌ Could not load YOLO. Make sure `ultralytics` is installed.")
            else:
                yolo = yolo_fast if 'yolov8n' in cfg['yolo_model'] else yolo_final

                st.markdown('<div class="section-title">⚙️ Processing</div>', unsafe_allow_html=True)
                prog_bar   = st.progress(0.0)
                status_txt = st.empty()
                log_box    = st.empty()

                try:
                    result = run_pipeline(
                        tmp_path, cfg, yolo, reid_model, dev, fp16,
                        progress_bar=prog_bar, status_text=status_txt, log_container=log_box)

                    prog_bar.progress(1.0)
                    status_txt.markdown("✅ **Pipeline complete!**")

                    st.session_state['result']   = result
                    st.session_state['cfg']      = cfg
                    st.session_state['vid_name'] = video_file.name

                    m = result['metrics']
                    st.success(f"Done! Detected **{m['total_persons']}** unique person(s) · "
                               f"{m['total_entries']} entries · {m['total_exits']} exits")

                    # Quick log preview
                    if result['log_lines']:
                        st.markdown('<div class="section-title">📋 Event Log Preview</div>',
                                    unsafe_allow_html=True)
                        log_html = '<div class="log-box">' + \
                                   '<br>'.join(result['log_lines'][-25:]) + '</div>'
                        st.markdown(log_html, unsafe_allow_html=True)

                except Exception as e:
                    st.error(f"Pipeline failed: {e}")
                finally:
                    os.unlink(tmp_path)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 2 — RESULTS & EVENTS
# ─────────────────────────────────────────────────────────────────────────────
with tab_results:
    if 'result' not in st.session_state:
        st.info("Run the pipeline first to see results here.")
    else:
        result = st.session_state['result']
        m      = result['metrics']
        cfg    = st.session_state['cfg']

        # ── KPI cards ──────────────────────────────────────────────────────────
        c1, c2, c3, c4, c5 = st.columns(5)
        with c1:
            st.markdown(f"""<div class="metric-card">
              <div class="label">Unique Persons</div>
              <div class="value">{m['total_persons']}</div></div>""", unsafe_allow_html=True)
        with c2:
            st.markdown(f"""<div class="metric-card">
              <div class="label">Entries</div>
              <div class="value green">{m['total_entries']}</div></div>""", unsafe_allow_html=True)
        with c3:
            st.markdown(f"""<div class="metric-card">
              <div class="label">Exits</div>
              <div class="value amber">{m['total_exits']}</div></div>""", unsafe_allow_html=True)
        with c4:
            st.markdown(f"""<div class="metric-card">
              <div class="label">Re-Identified</div>
              <div class="value purple">{m['revisit_persons']}</div></div>""", unsafe_allow_html=True)
        with c5:
            st.markdown(f"""<div class="metric-card">
              <div class="label">Pipeline Score</div>
              <div class="value">{result['score']:.2f}</div></div>""", unsafe_allow_html=True)

        st.markdown("---")

        col_l, col_r = st.columns(2, gap="large")

        with col_l:
            st.markdown('<div class="section-title">👥 Person Identity Table</div>',
                        unsafe_allow_html=True)
            db = result['database']
            rows = []
            for pid, rec in sorted(db.items()):
                rows.append({
                    'Person ID':    pid,
                    'First Frame':  rec.get('first_seen', '?'),
                    'Last Frame':   rec.get('last_seen',  '?'),
                    'Visit Count':  rec.get('visit_count', 0),
                    'Stored Embs':  len(rec.get('embeddings', [])),
                    'Re-ID':        '✅' if rec.get('visit_count', 0) > 1 else '—',
                })
            if rows:
                df_persons = pd.DataFrame(rows)
                st.dataframe(df_persons, use_container_width=True, hide_index=True,
                             height=min(400, 50 + 35 * len(rows)))
            else:
                st.warning("No persons detected.")

        with col_r:
            st.markdown('<div class="section-title">📋 Full Event Log</div>',
                        unsafe_allow_html=True)
            evs = result['event_log']
            if evs:
                df_ev = pd.DataFrame(evs)
                df_ev['event'] = df_ev['event'].apply(
                    lambda e: f"🟢 {e}" if e == 'ENTERED' else f"🔴 {e}")
                st.dataframe(df_ev, use_container_width=True, hide_index=True,
                             height=min(400, 50 + 35 * len(df_ev)))
            else:
                st.info("No entry/exit events were recorded.")

        # ── Config used ────────────────────────────────────────────────────────
        with st.expander("⚙️ Config used for this run", expanded=False):
            cfg_df = pd.DataFrame([{'Parameter': k, 'Value': v} for k, v in cfg.items()])
            st.dataframe(cfg_df, use_container_width=True, hide_index=True)


# ─────────────────────────────────────────────────────────────────────────────
# TAB 3 — ANALYTICS
# ─────────────────────────────────────────────────────────────────────────────
with tab_analytics:
    if 'result' not in st.session_state:
        st.info("Run the pipeline first to see analytics here.")
    else:
        result = st.session_state['result']
        db     = result['database']
        evs    = result['event_log']
        m      = result['metrics']
        cfg    = st.session_state['cfg']

        plt.style.use('dark_background')
        fig_bg = '#0f1117'
        ax_bg  = '#1c2333'
        grid_c = '#2a3045'
        blue   = '#4fc3f7'
        green  = '#66bb6a'
        red    = '#ef5350'
        amber  = '#ffa726'

        # ── Row 1: visits bar + entry/exit timeline ──────────────────────────
        col1, col2 = st.columns(2, gap="large")

        with col1:
            st.markdown('<div class="section-title">👤 Visits per Person</div>',
                        unsafe_allow_html=True)
            if db:
                pids_   = sorted(db.keys())
                visits_ = [db[p].get('visit_count', 0) for p in pids_]
                fig, ax = plt.subplots(figsize=(6, 3.2), facecolor=fig_bg)
                ax.set_facecolor(ax_bg)
                colors_ = [amber if v > 1 else blue for v in visits_]
                bars = ax.bar([str(p) for p in pids_], visits_, color=colors_,
                              width=0.6, zorder=3)
                ax.axhline(1, color=red, linestyle='--', linewidth=1.2, alpha=0.7,
                           label='single visit')
                for bar, v in zip(bars, visits_):
                    ax.text(bar.get_x() + bar.get_width()/2, bar.get_height() + 0.05,
                            str(v), ha='center', va='bottom', color='#e0e0e0', fontsize=9)
                ax.set_xlabel('Person ID', color='#7a8fb5', fontsize=9)
                ax.set_ylabel('Visit Count', color='#7a8fb5', fontsize=9)
                ax.set_title('Re-ID Visits per Person', color='#c0cce0', fontsize=10, pad=8)
                ax.tick_params(colors='#7a8fb5', labelsize=8)
                ax.spines[:].set_color(grid_c)
                ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
                ax.grid(axis='y', color=grid_c, linewidth=0.5, zorder=0)
                legend_patches = [
                    mpatches.Patch(color=amber, label='Re-identified (>1 visit)'),
                    mpatches.Patch(color=blue,  label='Single visit'),
                ]
                ax.legend(handles=legend_patches, loc='upper right',
                          fontsize=7.5, framealpha=0.3, facecolor=ax_bg)
                plt.tight_layout(pad=0.8)
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)

        with col2:
            st.markdown('<div class="section-title">⏱ Entry / Exit Timeline</div>',
                        unsafe_allow_html=True)
            if evs:
                df_ev = pd.DataFrame(evs)
                fig, ax = plt.subplots(figsize=(6, 3.2), facecolor=fig_bg)
                ax.set_facecolor(ax_bg)
                for evt, color, marker in [('ENTERED', green, '^'), ('EXITED', red, 'v')]:
                    sub = df_ev[df_ev['event'] == evt]
                    ax.scatter(sub['frame'], sub['person_id'],
                               label=evt, c=color, alpha=0.85, s=55, marker=marker, zorder=3)
                ax.set_xlabel('Frame', color='#7a8fb5', fontsize=9)
                ax.set_ylabel('Person ID', color='#7a8fb5', fontsize=9)
                ax.set_title('Entry & Exit Events over Time', color='#c0cce0', fontsize=10, pad=8)
                ax.tick_params(colors='#7a8fb5', labelsize=8)
                ax.spines[:].set_color(grid_c)
                ax.grid(color=grid_c, linewidth=0.5, zorder=0)
                ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
                ax.legend(fontsize=8, framealpha=0.3, facecolor=ax_bg)
                plt.tight_layout(pad=0.8)
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)
            else:
                st.info("No events recorded — check entry/exit zone settings.")

        # ── Row 2: detection lifespan + person active frames bar ─────────────
        col3, col4 = st.columns(2, gap="large")

        with col3:
            st.markdown('<div class="section-title">📏 Detection Lifespan</div>',
                        unsafe_allow_html=True)
            if db:
                lifespans = [(pid, db[pid].get('last_seen', 0) - db[pid].get('first_seen', 0))
                             for pid in sorted(db.keys())]
                pids_ls   = [str(l[0]) for l in lifespans]
                spans_ls  = [l[1] for l in lifespans]
                fig, ax = plt.subplots(figsize=(6, 3.2), facecolor=fig_bg)
                ax.set_facecolor(ax_bg)
                ax.barh(pids_ls, spans_ls, color='#7c4dff', height=0.55, zorder=3)
                ax.set_xlabel('Frames active', color='#7a8fb5', fontsize=9)
                ax.set_ylabel('Person ID', color='#7a8fb5', fontsize=9)
                ax.set_title('Person Track Lifespan (frames)', color='#c0cce0', fontsize=10, pad=8)
                ax.tick_params(colors='#7a8fb5', labelsize=8)
                ax.spines[:].set_color(grid_c)
                ax.grid(axis='x', color=grid_c, linewidth=0.5, zorder=0)
                plt.tight_layout(pad=0.8)
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)

        with col4:
            st.markdown('<div class="section-title">🗂 Embeddings Stored</div>',
                        unsafe_allow_html=True)
            if db:
                pids_e  = [str(p) for p in sorted(db.keys())]
                emb_cnt = [len(db[int(p)].get('embeddings', [])) for p in pids_e]
                fig, ax = plt.subplots(figsize=(6, 3.2), facecolor=fig_bg)
                ax.set_facecolor(ax_bg)
                ax.bar(pids_e, emb_cnt, color='#26a69a', width=0.6, zorder=3)
                ax.set_xlabel('Person ID', color='#7a8fb5', fontsize=9)
                ax.set_ylabel('Embedding count', color='#7a8fb5', fontsize=9)
                ax.set_title('Gallery Embeddings per Person', color='#c0cce0', fontsize=10, pad=8)
                ax.tick_params(colors='#7a8fb5', labelsize=8)
                ax.spines[:].set_color(grid_c)
                ax.grid(axis='y', color=grid_c, linewidth=0.5, zorder=0)
                ax.yaxis.set_major_locator(plt.MaxNLocator(integer=True))
                plt.tight_layout(pad=0.8)
                st.pyplot(fig, use_container_width=True)
                plt.close(fig)

        # ── Summary metrics row ────────────────────────────────────────────────
        st.markdown('<div class="section-title">📊 Pipeline Summary</div>', unsafe_allow_html=True)
        s1, s2, s3, s4 = st.columns(4)
        total_frames = m['frames_processed']
        proc_ratio   = (total_frames / max(total_frames, 1)) * 100
        re_id_rate   = m['revisit_persons'] / max(m['total_persons'], 1) * 100
        match_ratio  = min(m['total_entries'], m['total_exits']) / max(m['total_entries'], 1) * 100

        with s1:
            st.metric("Frames Processed",  f"{total_frames:,}")
        with s2:
            st.metric("Re-ID Rate",         f"{re_id_rate:.1f}%")
        with s3:
            st.metric("Entry/Exit Balance", f"{match_ratio:.1f}%")
        with s4:
            st.metric("Max Persons Cap",    cfg.get('max_persons_expected', 8))


# ─────────────────────────────────────────────────────────────────────────────
# TAB 4 — DOWNLOAD
# ─────────────────────────────────────────────────────────────────────────────
with tab_download:
    if 'result' not in st.session_state:
        st.info("Run the pipeline first to enable downloads.")
    else:
        result   = st.session_state['result']
        vid_name = st.session_state.get('vid_name', 'video')
        ts       = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')

        st.markdown('<div class="section-title">💾 Download Results</div>', unsafe_allow_html=True)
        st.markdown("""
        <p style="color:#7a8fb5; font-size:0.87rem; margin-bottom:18px;">
        All outputs from the most recent pipeline run are available below.
        </p>""", unsafe_allow_html=True)

        dl1, dl2, dl3 = st.columns(3, gap="large")

        # ── Excel ─────────────────────────────────────────────────────────────
        with dl1:
            st.markdown("""
            <div class="metric-card" style="text-align:left; padding:20px;">
              <div style="font-size:1.5rem; margin-bottom:8px;">📊</div>
              <div style="font-weight:700; color:#e0e0e0; margin-bottom:4px;">Excel Report</div>
              <div style="font-size:0.78rem; color:#7a8fb5; margin-bottom:16px;">
                Person table · Event log · Config · Summary
              </div>
            </div>""", unsafe_allow_html=True)
            try:
                excel_buf = results_to_excel(result)
                st.download_button(
                    label="⬇ Download Excel (.xlsx)",
                    data=excel_buf,
                    file_name=f"reid_report_{ts}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True)
            except Exception as e:
                st.error(f"Excel export failed: {e}")

        # ── CSV ────────────────────────────────────────────────────────────────
        with dl2:
            st.markdown("""
            <div class="metric-card" style="text-align:left; padding:20px;">
              <div style="font-size:1.5rem; margin-bottom:8px;">📄</div>
              <div style="font-weight:700; color:#e0e0e0; margin-bottom:4px;">CSV Export</div>
              <div style="font-size:0.78rem; color:#7a8fb5; margin-bottom:16px;">
                Event log · Person table combined
              </div>
            </div>""", unsafe_allow_html=True)
            db  = result['database']
            evs = result['event_log']
            rows = []
            for pid, rec in sorted(db.items()):
                rows.append({
                    'person_id':     pid,
                    'first_seen':    rec.get('first_seen', '?'),
                    'last_seen':     rec.get('last_seen',  '?'),
                    'visit_count':   rec.get('visit_count', 0),
                    'stored_embs':   len(rec.get('embeddings', [])),
                    're_identified': rec.get('visit_count', 0) > 1,
                })
            csv_persons = pd.DataFrame(rows).to_csv(index=False)
            st.download_button(
                label="⬇ Download Persons CSV",
                data=csv_persons,
                file_name=f"reid_persons_{ts}.csv",
                mime="text/csv",
                use_container_width=True)
            if evs:
                csv_events = pd.DataFrame(evs).to_csv(index=False)
                st.download_button(
                    label="⬇ Download Events CSV",
                    data=csv_events,
                    file_name=f"reid_events_{ts}.csv",
                    mime="text/csv",
                    use_container_width=True)

        # ── Output Video ───────────────────────────────────────────────────────
        with dl3:
            st.markdown("""
            <div class="metric-card" style="text-align:left; padding:20px;">
              <div style="font-size:1.5rem; margin-bottom:8px;">🎬</div>
              <div style="font-weight:700; color:#e0e0e0; margin-bottom:4px;">Output Video</div>
              <div style="font-size:0.78rem; color:#7a8fb5; margin-bottom:16px;">
                Annotated video with bounding boxes & IDs
              </div>
            </div>""", unsafe_allow_html=True)
            raw_path = result.get('output_path', '')
            if raw_path and os.path.exists(raw_path):
                # Try H.264 conversion
                with st.spinner("Encoding H.264…"):
                    dl_path = convert_to_h264(raw_path)
                with open(dl_path, 'rb') as f:
                    video_bytes = f.read()
                st.download_button(
                    label="⬇ Download Video (.mp4)",
                    data=video_bytes,
                    file_name=f"reid_output_{ts}.mp4",
                    mime="video/mp4",
                    use_container_width=True)
                st.info(f"File size: {len(video_bytes)/1024/1024:.1f} MB")
            else:
                st.warning("Output video not found.")

        # ── Config JSON ────────────────────────────────────────────────────────
        st.markdown('<div class="section-title">⚙️ Config & Metrics JSON</div>',
                    unsafe_allow_html=True)
        import json
        combined = {
            'run_timestamp': ts,
            'config':        result['cfg'],
            'metrics':       result['metrics'],
            'score':         result['score'],
        }
        st.download_button(
            label="⬇ Download Config + Metrics (.json)",
            data=json.dumps(combined, indent=2),
            file_name=f"reid_config_{ts}.json",
            mime="application/json")